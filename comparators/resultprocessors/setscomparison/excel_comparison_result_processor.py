from openpyxl import Workbook
from result_processor import AbstractResultProcessor


class ExcelResultProcessor(AbstractResultProcessor):
    NAME = 'excel-result-processor'

    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = 'Fuzzy set similarity result'

        self.results = []
        self.lastRow = 0

    def getName(self):
        return ExcelResultProcessor.NAME

    def processComparisonResult(self, resultComparatorResult, processingResultFile):
        print "Processing comparison result - saving result in " + processingResultFile
        numberOfSets = len(resultComparatorResult.listOfSets)
        lastRow = numberOfSets + 2
        resultMatrix = resultComparatorResult.resultMatrix
        resultsSum = 0
        for i in range(numberOfSets):
            for j in range(i, numberOfSets):
                self.sheet.cell(column=i + 1, row=j + 1, value=resultMatrix[i][j])
                resultsSum += resultMatrix[i][j]


        self.sheet.cell(column=1, row=lastRow, value="Average")
        avg = resultsSum / numberOfSets ** 2
        self.sheet.cell(column=2, row=lastRow, value=avg)

        self.workbook.save(processingResultFile)
