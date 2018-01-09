from openpyxl import Workbook

class ConnsoleResultProcessor(object):
    def processResult(self, result, metadata):
        print "Result: {}, config: {} ".format(result, metadata.config)


class ExcelResultProcessor(ConnsoleResultProcessor):

    def __init__(self, numberOfSets, workbookFile='sample.xlsx'):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = 'Fuzzy set similarity result'
        self.workbookFile = workbookFile

        self.results = []
        self.lastRow = 0

    def processResult(self, result, metadata):
        print "Result: {}, config: {}. Stored in cell: (row, col): ({}, {})".format(result, metadata.config,
                                                                                    metadata.setAindex,
                                                                                    metadata.setBindex)
        #rows and columns are 1-index base, that means that first column has index equal to 1
        #and setAindex and setBIndex are zero based,
        setBRow = metadata.setBindex + 1
        self.lastRow = setBRow + 1;
        self.results.append(result)
        self.sheet.cell(column=metadata.setAindex + 1, row=setBRow, value=result)

    def appendSummary(self):
        self.sheet.cell(column=1, row=self.lastRow, value="Average")
        avg = sum(self.results) / len(self.results)
        self.sheet.cell(column=2, row=self.lastRow, value=avg)

    def save(self):
        self.workbook.save(self.workbookFile)
