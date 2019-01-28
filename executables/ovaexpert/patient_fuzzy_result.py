from openpyxl import load_workbook


class PatientData(object):
    """Wrapper for list of patient data"""

    def __init__(self, listOfPatients):
        self.patients = listOfPatients

    def getPatientWithId(self, id):
        return filter(lambda patient: patient.id == id, self.patients)[0]


class Patient(object):
    def __init__(self, rowMap):
        self.allFields = rowMap
        self.headers = []
        self.id = ''

    def withHeaders(self, headers):
        self.headers = headers
        return self

    def withPatientId(self, patientId):
        self.id = patientId
        return self

    def formatHeaders(self):
        formattedHeader = ''
        for header in self.headers:
            formattedHeader += "%-25s " % header
        return formattedHeader

    def toString(self):
        patient = ''
        for header in self.headers:
            patient += "%-25s " % self.allFields[header]
        return patient

class FuzzyPatientResult(object):
    """Obiekt reprezentujacy wyliczenia roznymi metodami (LR1, LR2, Timmermann, Alcazar, SM)
      czy dana pacjentka jest zdrowa (MalignancyCharacter = 0) czy chora (MalignancyCharacter = 1)"""

    LR1_MIN_HEADER_INDEX = 3

    def __init__(self, rowMap):
        self.allFields = rowMap
        self.headers = []
        self.numericHeaders = []
        self.id = ''
        self.obscureLevel = 0
        self.malignancyCharacter = 0

    def withHeaders(self, headers):
        self.headers = headers
        self.numericHeaders = headers[FuzzyPatientResult.LR1_MIN_HEADER_INDEX:]
        return self

    def withPatientId(self, patientId):
        self.id = patientId
        return self

    def withObscureLevel(self, obscureLevel):
        self.obscureLevel = obscureLevel
        return self

    def withMalignancyCharacter(self, malignancyCharacter):
        self.malignancyCharacter = malignancyCharacter
        return self

    def getNumericValues(self):
        numericValues = []
        for numericValueHeader in self.numericHeaders:
            value = self.allFields[numericValueHeader]
            numericValues.append(float(value))
        return numericValues

    def toString(self):
        formattedNumericValues = ""
        for x in self.getNumericValues():
            formattedNumericValues = "{}\t{}".format(formattedNumericValues, "{0:.2f}".format(x))
        return "{}\t{}\t{}".format(self.id, self.malignancyCharacter, formattedNumericValues)


class PatientResultReader(object):
    PATIENT_ID_HEADER_INDEX = 0
    OBSCURE_LEVEL_HEADER_INDEX = 1
    MALIGNANCY_CHARACTER_LEVEL_INDEX = 2

    def readFuzzyPatientResults(self, resultFile, sheetName='fuzzy patients'):
        """
        Reads all fuzzy patient results from resultFile specified in constructor
        :return: list with  FuzzyPatientResult objects read from file
        """
        patientsWorkBook = load_workbook(filename=resultFile)
        patientsSheet = patientsWorkBook[sheetName]
        headers = PatientResultReader.__getHeaderNames__(patientsSheet)
        rowsAsMaps = PatientResultReader.__convertDataRows__(sheet=patientsSheet)

        return PatientResultReader.__toFuzzyResultPatientObject__(headers, rowsAsMaps)

    def readPatients(self, resultFile, sheetName='fuzzy patients'):
        patientsWorkBook = load_workbook(filename=resultFile)
        patientsSheet = patientsWorkBook[sheetName]
        headers = PatientResultReader.__getHeaderNames__(patientsSheet)
        rowsAsMaps = PatientResultReader.__convertDataRows__(sheet=patientsSheet)
        patients = PatientResultReader.__toPatientObject__(headers, rowsAsMaps)
        return PatientData(patients)

    @staticmethod
    def __toPatientObject__(headers, rowsAsMaps):
        patients = []
        for rowMap in rowsAsMaps:
            fuzzyPatient = Patient(rowMap) \
                .withHeaders(headers) \
                .withPatientId(rowMap[headers[PatientResultReader.PATIENT_ID_HEADER_INDEX]])
            patients.append(fuzzyPatient)

        return patients

    @staticmethod
    def __toFuzzyResultPatientObject__(headers, rowsAsMaps):
        fuzzyPatients = []
        for rowMap in rowsAsMaps:
            fuzzyPatient = FuzzyPatientResult(rowMap) \
                .withHeaders(headers) \
                .withPatientId(rowMap[headers[PatientResultReader.PATIENT_ID_HEADER_INDEX]]) \
                .withMalignancyCharacter(rowMap[headers[PatientResultReader.MALIGNANCY_CHARACTER_LEVEL_INDEX]]) \
                .withObscureLevel(rowMap[headers[PatientResultReader.OBSCURE_LEVEL_HEADER_INDEX]])
            fuzzyPatients.append(fuzzyPatient)

        return fuzzyPatients

    @staticmethod
    def __getHeaderNames__(sheet):
        HEADER_ROW_INDEX = 1
        max_col = sheet.max_column
        headers = []
        # Loop will print all columns name
        for i in range(1, max_col + 1):
            headerCell = sheet.cell(row=HEADER_ROW_INDEX, column=i)
            headers.append(headerCell.value)

        return headers

    @staticmethod
    def __getRowMap__(sheet, rowIndex):
        rowMap = {}
        headers = PatientResultReader.__getHeaderNames__(sheet=sheet)
        columnsSize = len(headers)
        for i in range(0, columnsSize):
            columnIndex = i + 1
            value = sheet.cell(row=rowIndex, column=columnIndex).value
            rowMap[headers[i]] = value

        return rowMap

    @staticmethod
    def __convertDataRows__(sheet):
        """Converts all data row of a sheet (skipping first row with headers) to maps with header name -> cell value 
        :param sheet: 
        :return: list with maps representing each row
        """
        rows = []
        numberOfRows = sheet.max_row
        for rowIndex in range(2, numberOfRows + 1):
            row = PatientResultReader.__getRowMap__(sheet=sheet, rowIndex=rowIndex)
            rows.append(row)
        return rows
